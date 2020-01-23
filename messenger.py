from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import os
import time
import random
import pyautogui
import re
from openpyxl import Workbook
from openpyxl import load_workbook

limit=1200
min=30
max=50

def msg_maker(name):
	path =os.getcwd()
	raw=''
	msg=[]
	with open (path +"/msg.txt",'r') as f:
		raw=f.read()  
	f.close()
	spinner_re=r"\{[a-zA-Z0-9\|\.\,\s]*\}"
	blocks=re.findall(spinner_re,raw)
	for block in blocks:
		choices=block.replace('{','').replace('}','').split('|')
		choice=random.choice(choices)
		raw=raw.replace(block,choice)
		if '%name%' in raw:
			raw=raw.replace('%name%',name)
	return raw



path=os.getcwd()
options=Options()
options.add_experimental_option('debuggerAddress','127.0.0.1:9000')

counter = 0
browser=webdriver.Chrome(path + '/chromedriver',options=options)
browser.implicitly_wait(100)
wb=load_workbook(path + '\details.xlsx')
sheet=wb.active
for i in range(2,sheet.max_row + 1):
	os.system('cls')
	print('----- PRESS CTRL + C TO EXIT -----')
	print('DAILY EMAIL SEARCHES LEFT ',limit-counter)
	url = sheet.cell(i,1).value
	if url is None:
		os.system('cls')
		print("ALL DONE")
		break

	print('CURRENT URL {}'.format(url))
	browser.get(url) 
	if '/in/' in url:
		name = browser.find_element_by_xpath('/html/body/div[5]/div[4]/div[3]/div/div/div/div/div[2]/main/div[1]/div/section/div[2]/div[2]/div[1]/ul[1]/li[1]')
		msg=msg_maker(name.text)
		msg_button=browser.find_element_by_xpath('/html/body/div[5]/div[4]/div[3]/div/div/div/div/div[2]/main/div[1]/div/section/div[2]/div[1]/div[2]/div/div/span[1]/div/button/span')
		msg_button.click()
		option=browser.find_element_by_xpath('/html/body/div[5]/div[4]/aside/div[2]/div[1]/form/footer/div[2]/div[2]/button')
		option.click()
		option2=browser.find_element_by_xpath('/html/body/div[5]/div[4]/aside/div[2]/div[1]/form/footer/div[2]/div[2]/div/fieldset/label[2]')
		option2.click()
		type=browser.find_element_by_xpath('/html/body/div[5]/div[4]/aside/div[2]/div[1]/form/div[3]/div/div[1]/div[1]/p')
		type.send_keys(msg)
		send=browser.find_element_by_xpath('/html/body/div[5]/div[4]/aside/div[2]/div[1]/form/footer/div[2]/div[1]/button')
		send.click()
		counter +=1
		
	elif '/sales/' in url:
		name=browser.find_element_by_xpath('/html/body/div[5]/main/div[1]/div[1]/div/div[1]/div[1]/div/dl/dt/span')
		msg=msg_maker(name.text)
		msg_button=browser.find_element_by_xpath('/html/body/div[5]/main/div[1]/div[1]/div/div[2]/div[1]/div[2]/button')
		msg_button.click()
		type=browser.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/section/div[2]/form[1]/section/textarea')
		type.send_keys(msg)
		send=browser.find_element_by_xpath('/html/body/div[5]/div[2]/div[2]/div/div/section/div[2]/form[1]/div/section/button[2]/span')
		send.click()
		counter +=1

	wait=random.randint(min,max)
	print("WAITING FOR {} seconds".format(wait))
	time.sleep(wait)


